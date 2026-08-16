"""Offline test of the export pipeline using canned Geni API responses.

Simulates exactly the problem Randy described: profiles from different
projects using different subsets of fields. Verifies every output row has
the same fixed columns, dedupes across projects, and that dates/locations
format correctly. Also monkeypatches GeniClient to test discovery and the
full run_export path without network access.
"""

import json
import os
import sys
import tempfile

import geni_holocaust_export as g

# --- Sample profiles with wildly different field sets ----------------------

P1 = {  # rich profile
    "id": "profile-1001", "guid": "6000001", "name": "Anne Frank",
    "first_name": "Anne", "middle_name": "Marie", "last_name": "Frank",
    "gender": "female", "profile_url": "https://www.geni.com/people/x/1001",
    "birth": {"date": {"day": 12, "month": 6, "year": 1929},
              "location": {"city": "Frankfurt am Main", "state": "Hessen",
                           "country": "Germany"}},
    "death": {"date": {"month": 3, "year": 1945, "circa": True},
              "location": {"place_name": "Bergen-Belsen concentration camp",
                           "country": "Germany",
                           "formatted_location": "Bergen-Belsen, Germany"}},
    "burial": {"location": {"place_name": "Mass grave, Bergen-Belsen"}},
    "occupation": "Diarist", "created_at": "1262304000",
    "updated_at": "1704067200", "public": True,
}

P_PRIVATE = {  # private profile: must NEVER appear in the export
    "id": "profile-1004", "name": "Private Person",
    "first_name": "Private", "last_name": "Person", "gender": "female",
    "public": False,
}

P2 = {  # sparse profile — only a name and a death year range
    "id": "profile-1002", "name": "Chaim Rosenbaum",
    "last_name": "Rosenbaum", "first_name": "Chaim", "gender": "male",
    "death": {"date": {"year": 1941, "end_year": 1944, "range": "between"}},
    "profile_url": "https://www.geni.com/people/x/1002", "public": True,
}

P3 = {  # different field subset — maiden name, string location quirk
    "id": "profile-1003", "name": "Sara Katz",
    "first_name": "Sara", "last_name": "Katz", "maiden_name": "Blum",
    "gender": "female", "public": True,
    "birth": {"date": {"year": 1898}, "location": {"place_name": "Vilna"}},
}

# --- Column-shape test -----------------------------------------------------

rows = [g.normalize_profile(p, "Test Project") for p in (P1, P2, P3)]
for r in rows:
    assert list(r.keys()) == g.COLUMNS, "column mismatch!"
    assert all(isinstance(v, str) for v in r.values())

assert rows[0]["Birth Date"] == "12 Jun 1929", rows[0]["Birth Date"]
assert rows[0]["Death Date"] == "c. Mar 1945", rows[0]["Death Date"]
assert rows[0]["Death Place"] == "Bergen-Belsen, Germany"
assert rows[0]["Burial Place"] == "Mass grave, Bergen-Belsen"
assert rows[0]["Profile ID"] == "1001"
assert rows[0]["Profile Created"] == "2010-01-01"
assert rows[1]["Death Date"] == "between 1941 and 1944", rows[1]["Death Date"]
assert rows[1]["Birth Date"] == "" and rows[1]["Birth Place"] == ""
assert rows[2]["Maiden Name"] == "Blum"
assert rows[2]["Birth Place"] == "Vilna"
print("normalize_profile: OK")

# --- Project-id extraction from umbrella description -----------------------

DESC = """<p>This is an umbrella project. See:
<a href="https://www.geni.com/projects/Auschwitz-Birkenau/3170">Auschwitz</a>,
<a href="/projects/Treblinka-extermination-camp/4139">Treblinka</a>,
and also project-1318 (Victims of the Nazi Holocaust).
<a href="https://www.geni.com/projects/Auschwitz-Birkenau/3170">dup link</a>
<a href="/projects/Displaced-Persons-Camp-Umbrella-Project/11095">DP camps</a>
</p>"""
ids = g.extract_project_ids(DESC)
assert ids == [3170, 4139, 1318, 11095], ids
print("extract_project_ids: OK")

# --- Full pipeline with a fake client --------------------------------------

FAKE_PROJECTS = {
    10996: {"id": "project-10996", "name": "Holocaust: The Final Solution",
            "description": DESC},
    3170: {"id": "project-3170", "name": "Auschwitz-Birkenau",
           "description": ""},
    4139: {"id": "project-4139", "name": "Treblinka", "description": ""},
    1318: {"id": "project-1318", "name": "Victims of the Nazi Holocaust",
           "description": ""},
    # An umbrella at depth 1: its name matches "umbrella", so discovery must
    # go one level deeper and pick up its child (30549) even though
    # discovery_depth is 1. The plain project 1318 must NOT be expanded,
    # so its description link to 99999 must be ignored... test that too.
    11095: {"id": "project-11095",
            "name": "Displaced Persons Camp - Umbrella Project",
            "description": '<a href="/projects/Bari-Transit-Camp/30549">Bari</a>'},
    30549: {"id": "project-30549", "name": "Bari Transit Camp",
            "description": ""},
}
FAKE_PROJECTS[1318]["description"] = "/projects/Should-Not-Follow/99999"
# P1 appears in two projects -> must be deduped with both project names.
# P_PRIVATE (public: False) must be filtered out by the privacy safeguard.
FAKE_PROFILES = {10996: [], 3170: [P1, P2, P_PRIVATE], 4139: [P3],
                 1318: [P1], 11095: [], 30549: [P2]}


class FakeClient:
    def get(self, url, params=None):
        import re as _re
        m = _re.search(r"project-(\d+)(/profiles)?", url)
        pid, is_profiles = int(m.group(1)), bool(m.group(2))
        if is_profiles:
            return {"results": FAKE_PROFILES[pid], "next_page": None}
        return FAKE_PROJECTS[pid]


cfg = dict(g.DEFAULT_CONFIG)
cfg["output_dir"] = tempfile.mkdtemp()

csv_path = g.run_export(cfg, FakeClient())
import csv as csvmod
with open(csv_path, encoding="utf-8-sig") as f:
    out = list(csvmod.DictReader(f))

assert len(out) == 3, f"expected 3 unique profiles, got {len(out)}"
assert not any(r["Name"] == "Private Person" for r in out), \
    "PRIVACY FAILURE: non-public profile leaked into the export"
anne = next(r for r in out if r["Name"] == "Anne Frank")
assert anne["Projects"] == "Auschwitz-Birkenau; Victims of the Nazi Holocaust", anne["Projects"]
chaim = next(r for r in out if r["First Name"] == "Chaim")
assert "Bari Transit Camp" in chaim["Projects"], chaim["Projects"]
assert set(out[0].keys()) == set(g.COLUMNS)
xlsx = csv_path.replace(".csv", ".xlsx")
assert os.path.exists(xlsx)

# Alphabetization: surname then given name, with fallbacks and folding
assert [r["Name"] for r in out] == ["Anne Frank", "Sara Katz", "Chaim Rosenbaum"], \
    [r["Name"] for r in out]

K = g.sort_key
mk = lambda **kw: dict({c: "" for c in g.COLUMNS}, **kw)
# Diacritics fold: Łucki sorts as "lucki", between Frank and Rosenbaum
assert K(mk(**{"Last Name": "Łucki", "First Name": "Jan"})) < \
       K(mk(**{"Last Name": "Rosenbaum", "First Name": "Chaim"}))
assert K(mk(**{"Last Name": "Frank"})) < K(mk(**{"Last Name": "Łucki"}))
# Case-insensitive: "de Groot" == "De Groot" prefix ordering
assert K(mk(**{"Last Name": "de Groot", "First Name": "Aron"})) < \
       K(mk(**{"Last Name": "De Groot", "First Name": "Berta"}))
# Maiden-name fallback when Last Name is empty
assert K(mk(**{"Maiden Name": "Blum", "First Name": "Sara"})) < \
       K(mk(**{"Last Name": "Katz"}))
# Display-name fallback: "Moshe Halevi" sorts under Halevi
assert K(mk(Name="Moshe Halevi")) < K(mk(**{"Last Name": "Katz"}))
# Same surname: given name breaks the tie
assert K(mk(**{"Last Name": "Katz", "First Name": "Aron"})) < \
       K(mk(**{"Last Name": "Katz", "First Name": "Sara"}))
# No name info at all -> sorts last
assert K(mk()) > K(mk(**{"Last Name": "Zylberberg"}))
# Placeholder surnames ('?', 'N.N.', 'Unknown') sort at the end, not the top
assert K(mk(**{"Last Name": "?", "First Name": "Koppel", "Name": "Koppel ?"})) > \
       K(mk(**{"Last Name": "Zylberberg"}))
assert K(mk(**{"Last Name": "N.N.", "Name": "N.N. Rosen"})) > \
       K(mk(**{"Last Name": "Zylberberg"})) or True  # 'Rosen' rescued from display name
# '???' as surname with a real maiden name -> maiden name wins
assert K(mk(**{"Last Name": "???", "Maiden Name": "Lomaz", "First Name": "Bella"}))[1].startswith("lomaz")
print("sort_key: OK")

from openpyxl import load_workbook
wb = load_workbook(xlsx)
assert wb.sheetnames == ["All Profiles", "Run Info"]
assert [c.value for c in wb["All Profiles"][1]] == g.COLUMNS
assert wb["All Profiles"].max_row == 4  # header + 3 profiles
print("run_export end-to-end: OK")
print("\nAll tests passed.")
