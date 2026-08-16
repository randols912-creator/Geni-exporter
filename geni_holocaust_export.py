#!/usr/bin/env python3
"""
Geni Holocaust Projects — monthly profile export.

Walks an umbrella project on Geni (default: "Holocaust: The Final Solution",
project 10996), discovers its sub-projects from the links in the project
description, downloads every profile in every project via the Geni API, and
writes ONE combined spreadsheet with a fixed, standardized set of columns —
regardless of which fields each profile actually uses.

Usage:
    python3 geni_holocaust_export.py --authorize   # one-time OAuth setup
    python3 geni_holocaust_export.py --projects-only  # preview discovered projects
    python3 geni_holocaust_export.py               # full export (for cron)

Options:
    --config PATH        Config file (default: geni_config.json next to script)
    --max-profiles N     Stop after N profiles per project (for testing)
    --csv-only           Skip the .xlsx and write only the .csv

Requires: python3, requests, openpyxl  (pip install requests openpyxl)
"""

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone

import requests

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

API_BASE = "https://www.geni.com/api"
OAUTH_AUTHORIZE_URL = "https://www.geni.com/platform/oauth/authorize"
OAUTH_TOKEN_URL = "https://www.geni.com/platform/oauth/request_token"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CONFIG_PATH = os.path.join(SCRIPT_DIR, "geni_config.json")

# Profile fields we ask the API for (keeps responses small and fast).
PROFILE_FIELDS = ",".join([
    "id", "guid", "name", "first_name", "middle_name", "last_name",
    "maiden_name", "gender", "birth", "death", "burial", "occupation",
    "is_alive", "public", "profile_url", "created_at", "updated_at",
])

# The fixed columns of the output spreadsheet, in order. Every run produces
# exactly these columns no matter which fields the profiles happen to use.
COLUMNS = [
    "Profile ID",
    "Profile URL",
    "Name",
    "First Name",
    "Middle Name",
    "Last Name",
    "Maiden Name",
    "Gender",
    "Birth Date",
    "Birth Place",
    "Death Date",
    "Death Place",
    "Burial Place",
    "Occupation",
    "Projects",
    "Profile Created",
    "Profile Updated",
]

DEFAULT_CONFIG = {
    "client_id": "YOUR_GENI_APP_KEY",
    "client_secret": "YOUR_GENI_APP_SECRET",
    "redirect_uri": "https://YOUR-REGISTERED-DOMAIN/callback",
    # Rate-limit stopgap: App 2133 is new and, until Geni approves it, capped
    # at the unapproved default of 1 request/10s. client_id_legacy/
    # client_secret_legacy hold the older, already-approved app's credentials
    # as a fallback (same redirect_uri works for both — no separate URL
    # needed). The legacy app must be authorized once, separately, with
    # --authorize-legacy (its tokens are kept in their own file — a token
    # issued under one app can't be reused under the other).
    #
    # app_mode:
    #   "auto"    (default) — start every run on the primary app; if it turns
    #             out to be severely rate-limited (observed X-API-Rate-Limit
    #             at/below AUTO_SWITCH_RATE_THRESHOLD, or repeated HTTP 429s),
    #             switch to the legacy app for the rest of THIS run and stay
    #             there (never switches back mid-run). The NEXT run starts on
    #             the primary app again, so once Geni approves it, auto mode
    #             picks that up on its own with nothing to change here.
    #   "primary" — always use the primary app; never auto-switch.
    #   "legacy"  — always use the legacy app from the start.
    "client_id_legacy": "",
    "client_secret_legacy": "",
    "token_file_legacy": os.path.join(SCRIPT_DIR, "geni_tokens_legacy.json"),
    "app_mode": "auto",
    "umbrella_project_id": 10996,
    "discovery_depth": 1,
    "extra_project_ids": [],
    "exclude_project_ids": [],
    "include_umbrella_profiles": True,
    # Privacy safeguard: only profiles marked public on Geni are exported.
    # Private profiles (visible to the authorizing account through family
    # or curator rights) are skipped and counted in the log.
    "public_only": True,
    "expand_name_patterns": ["umbrella", "portal"],
    "expand_project_ids": [],
    "max_discovery_depth": 4,
    "output_dir": os.path.join(SCRIPT_DIR, "exports"),
    "token_file": os.path.join(SCRIPT_DIR, "geni_tokens.json"),
    # Re-authorization by email: if the saved token stops working (before
    # or during a run), the script emails notify_email an authorization
    # link and waits for you to save the resulting URL into auth_code_file.
    "notify_email": "",
    "mail_method": "auto",   # "php" (website relay), "sendmail", or "smtp"
    "mail_from": "",
    "smtp": {"host": "", "port": 587, "username": "", "password": "", "from": ""},
    "auth_code_file": "geni_auth_code.txt",
    "auth_wait_hours": 24,
    # Fully automatic pickup: if callback_index.php is deployed at the
    # registered Callback URL, the script polls it and no manual step is
    # needed beyond clicking the link and approving.
    "auth_code_url": "",
    "auth_fetch_key": "",
}

log = logging.getLogger("geni_export")


# ---------------------------------------------------------------------------
# Config & tokens
# ---------------------------------------------------------------------------

def load_config(path):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2)
        sys.exit(
            f"No config file found. A template was written to {path}.\n"
            "Fill in client_id, client_secret and redirect_uri (from your Geni "
            "app registration), then run again with --authorize."
        )
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    cfg = dict(DEFAULT_CONFIG, **raw)
    # Backward compat with the earlier boolean toggle (use_legacy_app).
    if "app_mode" not in raw and "use_legacy_app" in raw:
        cfg["app_mode"] = "legacy" if raw["use_legacy_app"] else "auto"
    mode = cfg.get("app_mode", "auto")
    if mode not in ("auto", "primary", "legacy"):
        sys.exit(f"{path}: app_mode must be 'auto', 'primary', or 'legacy' "
                 f"(got {mode!r}).")
    if mode in ("auto", "legacy") and not (cfg.get("client_id_legacy")
                                           and cfg.get("client_secret_legacy")):
        if mode == "legacy":
            sys.exit(f"{path}: app_mode is 'legacy' but client_id_legacy / "
                     "client_secret_legacy are not both filled in.")
        log.info("app_mode is 'auto' but no client_id_legacy/"
                 "client_secret_legacy is set — this run will use the "
                 "primary app only, with no fallback available.")
    if cfg["client_id"] == "YOUR_GENI_APP_KEY":
        sys.exit(f"Please edit {path} and fill in your Geni app credentials.")
    return cfg


def load_tokens(cfg):
    try:
        with open(cfg["token_file"], encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def save_tokens(cfg, tokens):
    tokens["saved_at"] = int(time.time())
    tmp = cfg["token_file"] + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(tokens, f, indent=2)
    os.replace(tmp, cfg["token_file"])
    try:
        os.chmod(cfg["token_file"], 0o600)
    except OSError:
        pass


# ---------------------------------------------------------------------------
# OAuth
# ---------------------------------------------------------------------------

def auth_url(cfg):
    from urllib.parse import urlencode
    return OAUTH_AUTHORIZE_URL + "?" + urlencode({
        "client_id": cfg["client_id"],
        "redirect_uri": cfg["redirect_uri"],
    })


def extract_code(text):
    """Accept either a bare authorization code or the full redirect URL."""
    text = text.strip()
    if "code=" in text:
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(text).query)
        return qs.get("code", [text])[0]
    return text


def exchange_code(cfg, code):
    """Trade an authorization code for tokens. Raises RuntimeError with
    Geni's error text on failure."""
    resp = requests.post(OAUTH_TOKEN_URL, data={
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "redirect_uri": cfg["redirect_uri"],
        "code": code,
    }, timeout=30)
    if resp.status_code >= 400:
        raise RuntimeError(
            f"Geni rejected the code (HTTP {resp.status_code}): {resp.text}"
        )
    tokens = resp.json()
    if "access_token" not in tokens:
        raise RuntimeError(f"Unexpected response from Geni: {tokens}")
    save_tokens(cfg, tokens)
    return tokens


def authorize_interactive(cfg):
    """One-time interactive OAuth: prints the URL, user pastes the code."""
    print("\n1. Open this URL in a browser and sign in to Geni:\n")
    print("   " + auth_url(cfg))
    print(
        "\n2. After you approve, the browser is redirected to your redirect_uri"
        "\n   with ?code=... in the address bar. Copy that code value.\n"
    )
    code = extract_code(input("3. Paste the code (or the whole URL) here: "))
    try:
        exchange_code(cfg, code)
    except RuntimeError as e:
        sys.exit(
            f"{e}\nNote: codes expire after ~10 minutes and can only be used "
            "once — run --authorize again to get a fresh one."
        )
    print(f"\nTokens saved to {cfg['token_file']}. You can now run the export.")


def legacy_run_cfg(cfg):
    """A copy of cfg with client_id/client_secret/token_file swapped to the
    legacy app's values. Every existing OAuth/API helper reads those three
    keys, so passing this instead of cfg is all that's needed to act as the
    legacy app — no other function needs to know two apps exist."""
    if not (cfg.get("client_id_legacy") and cfg.get("client_secret_legacy")):
        sys.exit("client_id_legacy / client_secret_legacy are not set in the "
                 "config.")
    run_cfg = dict(cfg)
    run_cfg["client_id"] = cfg["client_id_legacy"]
    run_cfg["client_secret"] = cfg["client_secret_legacy"]
    run_cfg["token_file"] = cfg.get("token_file_legacy") \
        or (cfg["token_file"] + ".legacy")
    return run_cfg


def authorize_legacy_interactive(cfg):
    """One-time interactive OAuth for the legacy fallback app, kept in its
    own token file so it's ready whenever --ratecheck or a run in "auto"
    mode finds the primary app too rate-limited to use."""
    print("Authorizing the LEGACY Geni app (the rate-limit fallback).")
    authorize_interactive(legacy_run_cfg(cfg))


def send_notification_email(cfg, subject, body):
    """Send a notification email. Tries, in order (or just the configured
    mail_method): 'php' — relay through the callback page on your website,
    which sends with the web server's own mail() exactly like the Virtual
    Arnold daily digest (no credentials needed); 'sendmail' — the local
    sendmail binary, for when this script runs on the web host itself;
    'smtp' — classic SMTP with credentials from the config."""
    to = cfg.get("notify_email")
    if not to:
        log.warning("notify_email not set; cannot send: %s", subject)
        return False
    method = (cfg.get("mail_method") or "auto").lower()
    order = [method] if method in ("php", "sendmail", "smtp") else \
        ["php", "sendmail", "smtp"]

    for m in order:
        try:
            if m == "php" and cfg.get("auth_code_url") and \
                    cfg.get("auth_fetch_key"):
                resp = requests.post(cfg["auth_code_url"], data={
                    "notify": "1",
                    "fetch": cfg["auth_fetch_key"],
                    "subject": subject,
                    "body": body,
                }, timeout=30)
                if resp.status_code == 200 and resp.text.strip() == "OK":
                    log.info("Notification email sent via website relay: %s",
                             subject)
                    return True
                log.warning("Website mail relay answered %s: %s",
                            resp.status_code, resp.text.strip()[:120])

            elif m == "sendmail":
                sendmail = "/usr/sbin/sendmail"
                if os.path.exists(sendmail):
                    import subprocess
                    from email.message import EmailMessage
                    msg = EmailMessage()
                    msg["Subject"] = subject
                    msg["To"] = to
                    msg["From"] = cfg.get("mail_from") or f"geni-export@{os.uname().nodename}"
                    msg.set_content(body)
                    subprocess.run([sendmail, "-t", "-oi"],
                                   input=msg.as_bytes(), check=True,
                                   timeout=60)
                    log.info("Notification email sent via sendmail: %s",
                             subject)
                    return True

            elif m == "smtp":
                smtp_cfg = cfg.get("smtp") or {}
                if smtp_cfg.get("host") and smtp_cfg.get("username"):
                    import smtplib
                    from email.message import EmailMessage
                    msg = EmailMessage()
                    msg["Subject"] = subject
                    msg["From"] = smtp_cfg.get("from") or smtp_cfg["username"]
                    msg["To"] = to
                    msg.set_content(body)
                    port = int(smtp_cfg.get("port", 587))
                    if port == 465:
                        server = smtplib.SMTP_SSL(smtp_cfg["host"], port,
                                                  timeout=30)
                    else:
                        server = smtplib.SMTP(smtp_cfg["host"], port,
                                              timeout=30)
                        server.starttls()
                    with server:
                        server.login(smtp_cfg["username"],
                                     smtp_cfg["password"])
                        server.send_message(msg)
                    log.info("Notification email sent via SMTP: %s", subject)
                    return True
        except Exception as e:
            log.warning("Email method '%s' failed: %s", m, e)

    log.error("Could not send notification email: %s", subject)
    return False


def _auth_code_path(cfg):
    path = cfg.get("auth_code_file") or "geni_auth_code.txt"
    if not os.path.isabs(path):
        path = os.path.join(SCRIPT_DIR, path)
    return path


def poll_code_url(cfg):
    """Ask the deployed callback page whether an authorization code has
    arrived. Returns the code string, or None."""
    code_url = cfg.get("auth_code_url")
    key = cfg.get("auth_fetch_key")
    if not (code_url and key):
        return None
    try:
        resp = requests.get(code_url, params={"fetch": key}, timeout=30)
        code = resp.text.strip()
        if resp.status_code == 200 and code and len(code) < 200:
            return code
    except requests.RequestException as e:
        log.warning("Could not poll %s: %s", code_url, e)
    return None


def wait_for_reauthorization(cfg, reason):
    """Authorization is broken: email the user an authorization link, then
    wait (polling every minute, up to auth_wait_hours) until the code
    arrives — either automatically via the deployed callback page
    (auth_code_url) or manually via the auth code file. Returns tokens."""
    code_path = _auth_code_path(cfg)
    url = auth_url(cfg)
    automatic = bool(cfg.get("auth_code_url") and cfg.get("auth_fetch_key"))
    if automatic:
        body = (
            "The Geni Holocaust projects export needs to be re-authorized.\n"
            f"Reason: {reason}\n\n"
            "Open this link in your browser, sign in to Geni and approve:\n\n"
            f"   {url}\n\n"
            "That's it - the script will detect the authorization and resume "
            "automatically within a minute or two.\n"
        )
    else:
        body = (
            "The Geni Holocaust projects export needs to be re-authorized.\n"
            f"Reason: {reason}\n\n"
            "1. Open this link in your browser and sign in to Geni:\n\n"
            f"   {url}\n\n"
            "2. After you approve, the browser lands on a page whose address\n"
            "   contains ?code=... (the page itself may not load - that's fine).\n"
            "   Copy the ENTIRE address from the address bar.\n\n"
            "3. Paste it into a plain text file saved as:\n\n"
            f"   {code_path}\n\n"
            "The script checks for that file every minute and will resume "
            "automatically. Note the code expires about 10 minutes after you "
            "approve, so do step 3 right away; if it expires, just repeat from "
            "step 1.\n"
        )
    send_notification_email(
        cfg, "ACTION NEEDED: Geni export needs re-authorization", body
    )
    log.warning("Waiting for re-authorization. %s", reason)
    log.warning("Authorize at: %s", url)
    log.warning("Then save the redirect URL into: %s", code_path)

    deadline = time.time() + float(cfg.get("auth_wait_hours", 24)) * 3600
    while time.time() < deadline:
        code = None
        # Automatic path: the callback page on your website caught the code.
        code = poll_code_url(cfg)
        # Manual fallback: a code file dropped next to the script.
        if not code and os.path.exists(code_path):
            try:
                with open(code_path, encoding="utf-8") as f:
                    code = extract_code(f.read())
            finally:
                os.remove(code_path)
        if code:
            try:
                tokens = exchange_code(cfg, code)
                log.info("Re-authorization successful; resuming.")
                return tokens
            except RuntimeError as e:
                log.error("Authorization code did not work (%s); waiting "
                          "for a new one.", e)
                send_notification_email(
                    cfg, "Geni export: that code didn't work",
                    f"{e}\n\nPlease try again:\n\n{body}",
                )
        time.sleep(60)
    raise RuntimeError(f"Re-authorization timed out. Last reason: {reason}")


def refresh_access_token(cfg, tokens):
    resp = requests.post(OAUTH_TOKEN_URL, data={
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "grant_type": "refresh_token",
        "refresh_token": tokens["refresh_token"],
    }, timeout=30)
    resp.raise_for_status()
    new_tokens = resp.json()
    if "access_token" not in new_tokens:
        raise RuntimeError(f"Token refresh failed: {new_tokens}")
    # Geni may or may not rotate the refresh token; keep the old one if absent.
    if not new_tokens.get("refresh_token"):
        new_tokens["refresh_token"] = tokens["refresh_token"]
    save_tokens(cfg, new_tokens)
    return new_tokens


# ---------------------------------------------------------------------------
# API client with rate limiting
# ---------------------------------------------------------------------------

class GeniClient:
    """Small Geni API client that honors Geni's rate-limit headers
    (initial limit: 40 requests per 10 seconds) and refreshes the
    access token when it expires mid-run."""

    # Auto-switch thresholds ("auto" mode only): a fresh, unapproved Geni app
    # answers with X-API-Rate-Limit: 1 — comfortably below any real approved
    # tier — so treating anything at/below 5 as "not viable" has no risk of
    # mistaking a genuinely usable limit for a throttled one. The 429 count
    # is a second, independent trigger in case a low limit is never observed
    # directly (e.g. every response before the first 429 lacked the header).
    AUTO_SWITCH_RATE_THRESHOLD = 5
    AUTO_SWITCH_429_THRESHOLD = 3

    def __init__(self, cfg, tokens):
        self.cfg = cfg
        self.tokens = tokens
        self.session = requests.Session()
        self._rate_logged = False
        # Rate-limit monitoring (Geni sends X-API-Rate-Limit / -Window /
        # -Remaining on every response). We track these so a run can report
        # whether Geni is throttling this app.
        self._burst = False          # True during --ratecheck: don't self-throttle
        self.req_count = 0
        self.count_429 = 0
        self.rate_hits = 0           # times X-API-Rate-Remaining came back <= 0
        self.min_remaining = None
        self.obs_limit = None
        self.obs_window = None
        # Auto-failover to the legacy app when the primary proves too
        # rate-limited to use. "legacy" mode starts already-switched (cfg was
        # built by legacy_run_cfg before this client was constructed); "auto"
        # mode starts on the primary and may switch mid-run; "primary" mode
        # never switches.
        self.mode = cfg.get("app_mode", "auto")
        self.switched_to_legacy = (self.mode == "legacy")
        self.legacy_available = bool(cfg.get("client_id_legacy")
                                     and cfg.get("client_secret_legacy"))
        self._legacy_warned = False

    def _headers(self):
        return {"Authorization": f"Bearer {self.tokens['access_token']}"}

    def _switch_to_legacy(self, reason):
        """Switch every subsequent call to the legacy app. Returns True if the
        switch happened, False if it couldn't (not configured / not yet
        authorized) — in which case we keep going on the primary app rather
        than crash a cron run over a fallback that isn't ready."""
        if self.switched_to_legacy:
            return True
        if not self.legacy_available:
            if not self._legacy_warned:
                log.warning("Primary app appears rate-limited (%s), but no "
                            "legacy app is configured (client_id_legacy / "
                            "client_secret_legacy) — continuing on the "
                            "primary app.", reason)
                self._legacy_warned = True
            return False
        legacy_cfg = legacy_run_cfg(self.cfg)
        legacy_tokens = load_tokens(legacy_cfg)
        if not legacy_tokens:
            if not self._legacy_warned:
                log.warning("Primary app appears rate-limited (%s), but the "
                            "legacy app has not been authorized yet (run "
                            "--authorize-legacy) — continuing on the "
                            "primary app.", reason)
                self._legacy_warned = True
            return False
        try:
            legacy_tokens = refresh_access_token(legacy_cfg, legacy_tokens)
        except Exception as e:
            if not self._legacy_warned:
                log.warning("Primary app appears rate-limited (%s), and the "
                            "legacy app's saved token no longer works (%s) "
                            "— continuing on the primary app. Re-run "
                            "--authorize-legacy.", reason, e)
                self._legacy_warned = True
            return False
        log.warning("Switching to the LEGACY Geni app for the rest of this "
                    "run (%s). client_id starts %s...",
                    reason, legacy_cfg["client_id"][:8])
        self.cfg = legacy_cfg
        self.tokens = legacy_tokens
        self.switched_to_legacy = True
        self._rate_logged = False    # log the (different) new app's limit too
        return True

    def _observe_rate(self, resp):
        """Read and track the rate-limit headers on one response; warn loudly
        when X-API-Rate-Remaining is exhausted."""
        lim = resp.headers.get("X-API-Rate-Limit")
        win = resp.headers.get("X-API-Rate-Window")
        rem = resp.headers.get("X-API-Rate-Remaining")
        self.req_count += 1
        if lim is not None:
            self.obs_limit = lim
        if win is not None:
            self.obs_window = win
        if not self._rate_logged and lim:
            log.info("Geni API rate limit: %s requests per %s seconds",
                     lim, win or "?")
            self._rate_logged = True
        try:
            rem_i = int(rem) if rem is not None else None
        except ValueError:
            rem_i = None
        if rem_i is not None:
            if self.min_remaining is None or rem_i < self.min_remaining:
                self.min_remaining = rem_i
            if self._burst:
                log.info("  probe req %d: X-API-Rate-Remaining=%s "
                         "(X-API-Rate-Limit=%s, X-API-Rate-Window=%ss)",
                         self.req_count, rem, lim, win)
            if rem_i <= 0:
                self.rate_hits += 1
                log.warning("RATE LIMIT REACHED: X-API-Rate-Remaining=%s "
                            "(X-API-Rate-Limit=%s per %ss)", rem, lim, win)
        return rem_i

    def rate_summary(self):
        return ("Rate-limit summary: %d API requests; observed "
                "X-API-Rate-Limit=%s per %ss window; lowest "
                "X-API-Rate-Remaining seen = %s; times remaining<=0: %d; "
                "HTTP 429 responses: %d." % (
                    self.req_count, self.obs_limit, self.obs_window,
                    self.min_remaining, self.rate_hits, self.count_429))

    def get(self, url, params=None, _retried=False):
        # Network errors (timeouts, connection resets) are retried with
        # backoff instead of crashing a multi-hour run.
        last_err = None
        for attempt in range(6):
            try:
                resp = self.session.get(
                    url, params=params, headers=self._headers(), timeout=120
                )
                break
            except requests.RequestException as e:
                last_err = e
                wait = min(300, 15 * (2 ** attempt))
                log.warning("Network error (%s); retry %d/5 in %ds",
                            e.__class__.__name__, attempt + 1, wait)
                time.sleep(wait)
        else:
            raise RuntimeError(f"Giving up after repeated network errors: "
                               f"{last_err}")

        self._observe_rate(resp)

        # Access token expired mid-run -> refresh and retry. If the refresh
        # itself fails (token revoked/expired), fall back to emailing the
        # user for re-authorization and resume once they complete it.
        if resp.status_code == 401 and not _retried:
            log.info("Access token expired; refreshing...")
            try:
                self.tokens = refresh_access_token(self.cfg, self.tokens)
            except Exception as e:
                self.tokens = wait_for_reauthorization(
                    self.cfg, f"Token refresh failed mid-run ({e})."
                )
            return self.get(url, params=params, _retried=True)

        # Explicit throttling -> wait and retry (or, in "auto" mode, switch to
        # the legacy app and retry THIS call there instead of sleeping through
        # a limit that clearly isn't workable).
        if resp.status_code == 429:
            self.count_429 += 1
            if self._burst:
                return None          # probe: stop instead of self-throttling
            if self.mode == "auto" and not self.switched_to_legacy \
                    and self.count_429 >= self.AUTO_SWITCH_429_THRESHOLD \
                    and self._switch_to_legacy(
                        f"{self.count_429} consecutive HTTP 429 responses"):
                return self.get(url, params=params, _retried=False)
            wait = int(resp.headers.get("Retry-After", 11))
            log.warning("HTTP 429 (rate limited); sleeping %ss", wait)
            time.sleep(wait)
            return self.get(url, params=params, _retried=_retried)

        resp.raise_for_status()

        # A successful response that also reveals a too-low rate limit: keep
        # and return this result (no need to re-fetch what we already have),
        # but switch so every call FROM HERE ON uses the legacy app instead of
        # continuing to crawl at an unworkable rate.
        if self.mode == "auto" and not self.switched_to_legacy \
                and not self._burst and self.obs_limit is not None:
            try:
                if int(self.obs_limit) <= self.AUTO_SWITCH_RATE_THRESHOLD:
                    self._switch_to_legacy(
                        f"observed X-API-Rate-Limit={self.obs_limit}")
            except ValueError:
                pass

        # Proactive throttling from Geni's rate-limit headers. Only slow
        # down when Geni explicitly says we're nearly out of quota — a
        # missing header must NOT cause a sleep.
        remaining = resp.headers.get("X-API-Rate-Remaining")
        window = resp.headers.get("X-API-Rate-Window", "10")
        try:
            if remaining is not None and int(remaining) <= 1 and not self._burst:
                time.sleep(int(window) + 1)
        except ValueError:
            pass

        return resp.json()


# ---------------------------------------------------------------------------
# Project discovery
# ---------------------------------------------------------------------------

PROJECT_LINK_RE = re.compile(
    r"(?:geni\.com)?/projects/[^\s\"'<>)]*?/(\d+)|project-(\d+)", re.I
)


def extract_project_ids(text):
    """Pull Geni project ids out of free text / HTML (a project description)."""
    ids = []
    for m in PROJECT_LINK_RE.finditer(text or ""):
        pid = int(m.group(1) or m.group(2))
        if pid not in ids:
            ids.append(pid)
    return ids


def should_expand(pid, name, depth, cfg):
    """Should we follow the project links in this project's description?

    Always expand up to discovery_depth. Beyond that, expand projects that
    are themselves umbrellas — matched by name (e.g. 'Umbrella', 'Portal')
    or listed explicitly in expand_project_ids — up to max_discovery_depth.
    """
    if depth >= int(cfg.get("max_discovery_depth", 4)):
        return False
    if depth < int(cfg["discovery_depth"]):
        return True
    if int(pid) in {int(x) for x in cfg.get("expand_project_ids", [])}:
        return True
    lowered = (name or "").lower()
    return any(p.lower() in lowered for p in cfg.get("expand_name_patterns", []))


def discover_projects(client, cfg):
    """Start from the umbrella project and collect sub-projects listed in its
    description (recursing to discovery_depth). Returns {id: name}."""
    umbrella_id = int(cfg["umbrella_project_id"])
    exclude = {int(x) for x in cfg["exclude_project_ids"]}
    projects = {}          # id -> name
    seen = set()
    frontier = [(umbrella_id, 0)]

    while frontier:
        pid, depth = frontier.pop(0)
        if pid in seen or pid in exclude:
            continue
        seen.add(pid)
        try:
            data = client.get(
                f"{API_BASE}/project-{pid}",
                params={"fields": "id,name,description"},
            )
        except requests.HTTPError as e:
            log.warning("Skipping project %s (HTTP error: %s)", pid, e)
            continue

        name = data.get("name", f"project-{pid}")
        is_umbrella = pid == umbrella_id
        if not is_umbrella or cfg["include_umbrella_profiles"]:
            projects[pid] = name
        log.info("Discovered project %s: %s (depth %d)", pid, name, depth)

        if should_expand(pid, name, depth, cfg):
            for child in extract_project_ids(data.get("description", "")):
                if child not in seen:
                    frontier.append((child, depth + 1))

    for pid in cfg["extra_project_ids"]:
        pid = int(pid)
        if pid not in projects and pid not in exclude:
            try:
                data = client.get(
                    f"{API_BASE}/project-{pid}", params={"fields": "id,name"}
                )
                projects[pid] = data.get("name", f"project-{pid}")
                log.info("Added extra project %s: %s", pid, projects[pid])
            except requests.HTTPError as e:
                log.warning("Skipping extra project %s (%s)", pid, e)

    return projects


# ---------------------------------------------------------------------------
# Profile normalization
# ---------------------------------------------------------------------------

def format_date(d):
    """Geni structured date -> readable string. Handles partial dates,
    'circa', and ranges. Examples: '1898', 'c. May 1898', '12 May 1898',
    'between 1940 and 1945'."""
    if not d:
        return ""
    if isinstance(d, str):
        return d

    def part(day, month, year):
        months = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        bits = []
        if day:
            bits.append(str(day))
        if month:
            try:
                bits.append(months[int(month)])
            except (ValueError, IndexError):
                bits.append(str(month))
        if year:
            bits.append(str(year))
        return " ".join(bits)

    start = part(d.get("day"), d.get("month"), d.get("year"))
    end = part(d.get("end_day"), d.get("end_month"), d.get("end_year"))

    if d.get("range") == "between" and end:
        text = f"between {start} and {end}"
    elif d.get("range") in ("before", "after") and start:
        text = f"{d['range']} {start}"
    else:
        text = start

    if d.get("circa") and text:
        text = f"c. {text}"
    return text


def format_location(loc):
    """Geni structured location -> single readable string."""
    if not loc:
        return ""
    if isinstance(loc, str):
        return loc
    if loc.get("formatted_location"):
        return loc["formatted_location"]
    if loc.get("place_name") and not any(
        loc.get(k) for k in ("city", "state", "country")
    ):
        return loc["place_name"]
    parts = [loc.get(k) for k in
             ("place_name", "city", "county", "state", "country")]
    return ", ".join(p for p in parts if p)


def format_timestamp(ts):
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
    except (ValueError, OverflowError, OSError):
        return str(ts)


def event(profile, key):
    """Return (date, location) strings for a profile event like birth/death."""
    ev = profile.get(key) or {}
    if not isinstance(ev, dict):
        return "", ""
    return format_date(ev.get("date")), format_location(ev.get("location"))


def normalize_profile(profile, project_name):
    """Map a raw Geni profile (whatever fields it has) onto the fixed
    COLUMNS. Missing fields become empty strings — this is what makes every
    export identical in shape."""
    birth_date, birth_place = event(profile, "birth")
    death_date, death_place = event(profile, "death")
    _, burial_place = event(profile, "burial")

    occupation = profile.get("occupation", "")
    if isinstance(occupation, dict):  # occasionally localized
        occupation = next(iter(occupation.values()), "")

    return {
        "Profile ID": str(profile.get("id", "")).replace("profile-", ""),
        "Profile URL": profile.get("profile_url") or profile.get("url", ""),
        "Name": profile.get("name", ""),
        "First Name": profile.get("first_name", ""),
        "Middle Name": profile.get("middle_name", ""),
        "Last Name": profile.get("last_name", ""),
        "Maiden Name": profile.get("maiden_name", ""),
        "Gender": (profile.get("gender") or "").capitalize(),
        "Birth Date": birth_date,
        "Birth Place": birth_place,
        "Death Date": death_date,
        "Death Place": death_place,
        "Burial Place": burial_place,
        "Occupation": occupation,
        "Projects": project_name,
        "Profile Created": format_timestamp(profile.get("created_at")),
        "Profile Updated": format_timestamp(profile.get("updated_at")),
    }


# ---------------------------------------------------------------------------
# Download & export
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Run state (checkpoint/resume)
# ---------------------------------------------------------------------------

def _state_path(cfg):
    return os.path.join(cfg["output_dir"], "run_state.json")


def load_state(cfg):
    try:
        with open(_state_path(cfg), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return {}


def save_state(cfg, state):
    os.makedirs(cfg["output_dir"], exist_ok=True)
    tmp = _state_path(cfg) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f)
    os.replace(tmp, _state_path(cfg))


def profile_key(profile):
    """Stable dedupe key, identical whether the row comes fresh from the
    API or is reloaded from the partial-run CSV."""
    return (str(profile.get("id", "")).replace("profile-", "")
            or str(profile.get("guid", "")) or repr(profile))


def _fold(s):
    """Sorting key normalization: strip accents/diacritics and case so that
    'Łódź', 'Lodz' and 'LODZ' sort together."""
    import unicodedata
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # A few letters NFKD can't decompose:
    s = s.translate(str.maketrans("łŁđĐøØßæÆœŒ", "lLdDoOsaAoO"))
    return s.casefold().strip()


def _is_placeholder(s):
    """True for stand-in 'names' like '?', '???', 'N.N.', 'Unknown'."""
    t = _fold(s).strip(" .?*-_'\"()")
    return t in ("", "nn", "n n", "unknown", "unbekannt", "nieznany",
                 "nezinams", "x", "xx", "xxx")


def sort_key(row):
    """Alphabetize by surname, then given name. Geni name fields are
    inconsistent, so fall back sensibly: surname = Last Name, else Maiden
    Name, else the last word of the display Name — ignoring placeholder
    values like '?' or 'N.N.'. Profiles where no real surname can be
    determined sort at the end instead of the top."""
    surname = ""
    for cand in (row["Last Name"], row["Maiden Name"]):
        if cand and not _is_placeholder(cand):
            surname = cand
            break
    given = row["First Name"]
    if not surname and row["Name"]:
        parts = [p for p in row["Name"].split() if not _is_placeholder(p)]
        if len(parts) >= 2:
            surname = parts[-1]
            given = given or " ".join(parts[:-1])
    return (0 if surname else 1, _fold(surname), _fold(given),
            _fold(row["Middle Name"]), _fold(row["Name"]))


def fetch_project_profiles(client, pid, max_profiles=None, start_page=1,
                           on_checkpoint=None):
    """Yield raw profile dicts for one project, following pagination.
    start_page resumes mid-project; on_checkpoint(page) is called every 20
    pages so the caller can persist progress."""
    url = f"{API_BASE}/project-{pid}/profiles"
    params = {"fields": PROFILE_FIELDS, "per_page": 50, "page": start_page}
    count = 0
    pages = start_page - 1
    while url:
        data = client.get(url, params=params)
        pages += 1
        if pages == start_page and data.get("total_count"):
            log.info("    project %s has %s profiles", pid,
                     data["total_count"])
        if pages % 20 == 0:
            log.info("    ...page %d (%d profiles this run)", pages, count)
            if on_checkpoint:
                on_checkpoint(pages)
        for profile in data.get("results", []):
            yield profile
            count += 1
            if max_profiles and count >= max_profiles:
                return
        url = data.get("next_page")
        # Re-attach the fields selection in case next_page doesn't carry it.
        params = None if (url and "fields=" in url) else {"fields": PROFILE_FIELDS}


def run_export(cfg, client, max_profiles=None, csv_only=False):
    os.makedirs(cfg["output_dir"], exist_ok=True)
    partial_path = os.path.join(cfg["output_dir"], "partial_run.csv")
    month = datetime.now().strftime("%Y-%m")

    # Resume an interrupted run for this month if there is one; otherwise
    # start fresh (rediscovering the project list).
    state = load_state(cfg)
    rows = {}  # profile key -> normalized row (deduped across projects)
    resumed = False
    if (state.get("month") == month and not state.get("done")
            and state.get("projects")):
        projects = {int(k): v for k, v in state["projects"].items()}
        # Recover rows from the partial snapshot — or, if the crash
        # happened during finalization (partial already cleaned up), from
        # this month's final CSV.
        src = partial_path if os.path.exists(partial_path) else None
        if not src:
            import glob
            cands = sorted(
                glob.glob(os.path.join(cfg["output_dir"],
                                       "geni_holocaust_export_*.csv")),
                key=os.path.getmtime)
            if cands:
                src = cands[-1]
                log.info("Partial snapshot missing; recovering rows "
                         "from %s", src)
        if src:
            with open(src, encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    rows[row["Profile ID"] or repr(row)] = row
        if rows or not state.get("completed"):
            resumed = True
            log.info("RESUMING this month's run: %d of %d projects done, "
                     "%d profiles already downloaded",
                     len(state.get("completed", [])), len(projects),
                     len(rows))
        else:
            log.warning("Checkpoint rows unrecoverable; restarting this "
                        "month's run from scratch")
    if not resumed:
        projects = discover_projects(client, cfg)
        state = {"month": month, "done": False, "completed": [],
                 "in_progress": {}, "projects": projects}
        save_state(cfg, state)
        log.info("Exporting %d projects", len(projects))

    completed = {int(x) for x in state.get("completed", [])}
    public_only = bool(cfg.get("public_only", True))
    skipped_private = [0]

    def add_profile(profile, pname):
        # Privacy safeguard: never export profiles that are not public on
        # Geni, even if the authorizing account is allowed to see them.
        if public_only and profile.get("public") is not True:
            skipped_private[0] += 1
            return
        key = profile_key(profile)
        if key in rows:
            existing = rows[key]["Projects"]
            if pname not in existing.split("; "):
                rows[key]["Projects"] = existing + "; " + pname
        else:
            rows[key] = normalize_profile(profile, pname)

    done = len(completed)
    for pid, pname in projects.items():
        if pid in completed:
            continue
        in_prog = state.get("in_progress") or {}
        start_page = in_prog.get("page", 1) if in_prog.get("pid") == pid else 1
        if start_page > 1:
            log.info("  resuming %s from page %d", pname, start_page)

        def checkpoint(page, _pid=pid):
            state["in_progress"] = {"pid": _pid, "page": page}
            write_csv(partial_path, rows.values())
            save_state(cfg, state)

        n = 0
        for profile in fetch_project_profiles(client, pid, max_profiles,
                                              start_page=start_page,
                                              on_checkpoint=checkpoint):
            add_profile(profile, pname)
            n += 1
        done += 1
        log.info("  [%d/%d] %s: %d profiles", done, len(projects), pname, n)
        # Checkpoint: this project is finished; snapshot rows to disk so an
        # interrupted run resumes here instead of starting over.
        state["completed"].append(pid)
        state["in_progress"] = {}
        write_csv(partial_path, rows.values())
        save_state(cfg, state)

    all_rows = sorted(rows.values(), key=sort_key)
    log.info("Total unique profiles: %d", len(all_rows))
    if skipped_private[0]:
        log.info("Skipped %d non-public profiles (public_only safeguard)",
                 skipped_private[0])

    stamp = datetime.now().strftime("%Y-%m-%d")
    base = os.path.join(cfg["output_dir"], f"geni_holocaust_export_{stamp}")

    csv_path = base + ".csv"
    write_csv(csv_path, all_rows)
    log.info("Wrote %s", csv_path)
    if os.path.exists(partial_path):
        os.remove(partial_path)

    xlsx_path = None
    if not csv_only:
        try:
            xlsx_path = write_xlsx(base + ".xlsx", all_rows, projects)
            log.info("Wrote %s", xlsx_path)
        except Exception as e:
            # The CSV is complete; never let a spreadsheet-build problem
            # block finishing the month and sending the email.
            log.error("Could not build the .xlsx (%s); finishing with "
                      "the .csv only.", e)

    state["done"] = True
    state["in_progress"] = {}
    save_state(cfg, state)

    paths = [p for p in [xlsx_path, csv_path] if p]
    if cfg.get("auth_code_url") and cfg.get("auth_fetch_key"):
        from urllib.parse import urlencode
        links = "\n".join(
            "   " + cfg["auth_code_url"].rstrip("/") + "/?" + urlencode(
                {"download": os.path.basename(p),
                 "fetch": cfg["auth_fetch_key"]})
            for p in paths)
        where = f"Download the spreadsheet here:\n\n{links}"
    else:
        where = ("Files written on the server:\n"
                 + "\n".join("   " + p for p in paths)
                 + "\n\nDownload them via cPanel File Manager.")
    send_notification_email(
        cfg,
        f"Geni Holocaust export finished: {len(all_rows):,} profiles",
        f"The monthly Geni Holocaust projects export completed successfully.\n\n"
        f"   Projects covered: {len(projects)}\n"
        f"   Unique profiles:  {len(all_rows):,}\n\n"
        f"{where}\n",
    )
    return csv_path


def write_csv(path, rows_iterable):
    tmp = path + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows_iterable)
    os.replace(tmp, path)


_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _xlsx_safe(value):
    """Strip control characters that are illegal in Excel files (they occur
    occasionally in Geni project names and profile data)."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def write_xlsx(path, all_rows, projects):
    """Streaming (write-only) workbook build: memory stays flat no matter
    how many rows, so a 300k-profile export can't be killed for RAM."""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=True)
    bold = Font(bold=True)

    ws = wb.create_sheet("All Profiles")
    widths = {"Profile URL": 45, "Name": 30, "Birth Place": 30,
              "Death Place": 30, "Burial Place": 30, "Projects": 40}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNS))}{len(all_rows) + 1}"
    )
    header = []
    for col in COLUMNS:
        c = WriteOnlyCell(ws, value=col)
        c.font = bold
        header.append(c)
    ws.append(header)
    for row in all_rows:
        ws.append([_xlsx_safe(row[c]) for c in COLUMNS])

    meta = wb.create_sheet("Run Info")
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 60
    meta.append(["Export date", datetime.now().strftime("%Y-%m-%d %H:%M")])
    meta.append(["Total unique profiles", len(all_rows)])
    meta.append([])
    h1 = WriteOnlyCell(meta, value="Project ID")
    h2 = WriteOnlyCell(meta, value="Project name")
    h1.font = h2.font = bold
    meta.append([h1, h2])
    for pid, pname in sorted(projects.items()):
        meta.append([pid, _xlsx_safe(pname)])

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_ratecheck(cfg, burst=150):
    """Deliberately push the API to its rate limit and report what the
    X-API-Rate-* headers show. Fires light, read-only calls (project fields=id)
    as fast as the network allows, without self-throttling, and stops as soon
    as X-API-Rate-Remaining hits 0 or a 429 comes back."""
    tokens = load_tokens(cfg)
    if not tokens:
        log.error("No Geni tokens found; run --authorize first.")
        return
    client = GeniClient(cfg, tokens)
    pid = int(cfg["umbrella_project_id"])
    client._burst = True
    client.mode = "primary"    # never let a probe run auto-switch mid-measurement
    log.info("Rate-limit probe: sending up to %d light read-only calls "
             "(GET project-%s?fields=id) with no self-throttling, watching "
             "X-API-Rate-Remaining ...", burst, pid)
    for i in range(burst):
        try:
            client.get(f"{API_BASE}/project-{pid}", params={"fields": "id"})
        except Exception as e:            # noqa: BLE001
            log.warning("  probe request %d failed: %s", i + 1, e)
        if client.rate_hits or client.count_429:
            break
    log.info(client.rate_summary())
    if client.rate_hits or client.count_429:
        log.info("CONFIRMED: this app reaches Geni's rate limit "
                 "(X-API-Rate-Remaining hit 0 and/or HTTP 429). Send the "
                 "summary line above to api@geni.com.")
    else:
        log.info("Did not reach the limit within %d calls - the cap may be "
                 "higher, or the window reset before we could fill it. Re-run "
                 "with a larger --ratecheck N.", burst)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--authorize", action="store_true",
                        help="Run one-time interactive OAuth setup")
    parser.add_argument("--authorize-legacy", action="store_true",
                        help="One-time OAuth setup for the legacy fallback "
                             "app (client_id_legacy/client_secret_legacy), "
                             "kept in its own token file for app_mode=auto/"
                             "legacy to use")
    parser.add_argument("--projects-only", action="store_true",
                        help="Only list the projects that would be exported")
    parser.add_argument("--max-profiles", type=int, default=None,
                        help="Limit profiles per project (testing)")
    parser.add_argument("--csv-only", action="store_true")
    parser.add_argument("--cron", action="store_true",
                        help="Watchdog mode for an hourly cron: exit "
                             "quietly if this month's export is already "
                             "done or another instance is running; "
                             "otherwise start/resume the export.")
    parser.add_argument("--ratecheck", nargs="?", type=int, const=150,
                        metavar="N",
                        help="probe the API rate limit (N light calls, "
                             "default 150) and report the X-API-Rate-* headers")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(os.path.join(SCRIPT_DIR, "geni_export.log")),
        ],
    )

    cfg = load_config(args.config)

    if args.authorize:
        authorize_interactive(cfg)
        return

    if args.authorize_legacy:
        authorize_legacy_interactive(cfg)
        return

    if args.ratecheck is not None:
        run_ratecheck(cfg, burst=args.ratecheck)
        return

    pid_path = os.path.join(SCRIPT_DIR, "geni_export.pid")
    if args.cron:
        # Already finished this month? Nothing to do.
        state = load_state(cfg)
        if state.get("month") == datetime.now().strftime("%Y-%m") \
                and state.get("done"):
            return
        # Another instance still running? Leave it alone. The PID must
        # belong to a live process that is actually THIS script — process
        # IDs get recycled, and a stale pidfile must never block the
        # watchdog forever.
        try:
            with open(pid_path) as f:
                other = int(f.read().strip())
            with open(f"/proc/{other}/cmdline", "rb") as f:
                cmdline = f.read().decode("utf-8", "ignore")
            if "geni_holocaust" in cmdline:
                return
            log.info("Ignoring stale pidfile (PID %d is not this script).",
                     other)
        except (OSError, ValueError):
            pass  # no live instance; fall through and start/resume
        log.info("Cron watchdog: starting/resuming this month's export.")
    with open(pid_path, "w") as f:
        f.write(str(os.getpid()))

    # app_mode "legacy" starts directly on the legacy app; "auto" and
    # "primary" start on the primary app ("auto" may still switch mid-run —
    # see GeniClient._switch_to_legacy).
    mode = cfg.get("app_mode", "auto")
    run_cfg = legacy_run_cfg(cfg) if mode == "legacy" else cfg
    if mode == "legacy":
        log.info("app_mode=legacy: using the LEGACY Geni app for this run "
                 "(client_id starts %s...)", run_cfg["client_id"][:8])

    # Pre-run authorization check: always start with a fresh access token
    # (they expire daily, so a monthly cron run always finds the old one
    # expired). If there are no tokens or the refresh fails, email the user
    # an authorization link and wait for them to complete it.
    tokens = load_tokens(run_cfg)
    if not tokens:
        if run_cfg.get("notify_email"):
            tokens = wait_for_reauthorization(run_cfg, "No saved tokens found.")
        else:
            sys.exit("No saved tokens. Run with --authorize"
                     + (" --authorize-legacy" if mode == "legacy" else "")
                     + " first.")
    else:
        try:
            tokens = refresh_access_token(run_cfg, tokens)
        except Exception as e:
            tokens = wait_for_reauthorization(
                run_cfg, f"Stored token no longer works ({e})."
            )
    client = GeniClient(run_cfg, tokens)

    if args.projects_only:
        projects = discover_projects(client, cfg)
        print(f"\n{len(projects)} projects would be exported:")
        for pid, pname in sorted(projects.items()):
            print(f"  {pid:>8}  {pname}")
        return

    run_export(cfg, client, max_profiles=args.max_profiles,
               csv_only=args.csv_only)


if __name__ == "__main__":
    main()
